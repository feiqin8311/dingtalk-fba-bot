from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from fba_alert.alerts import build_listing_contact_map, classify_record, parse_summary_items
from fba_alert.lingxing import (
    InventorySnapshot,
    aggregate_inventory_snapshot,
    is_source_list_rate_limited_response,
)
from fba_alert.report import REPORT_HEADERS, build_report_rows, export_alert_report, export_scoped_alert_report
from openpyxl import load_workbook
from tests.factories import make_summary_item


class SummaryDailySalesTests(unittest.TestCase):
    def test_classify_record_uses_libraton_jp_jp_a_level_thresholds(self) -> None:
        item = {
            "basic_info": {
                "asin": "B1457A",
                "hash_id": "hash-1457-a",
                "sid": "1457",
                "node_type": 1,
                "msku_fnsku_list": [{"msku": "MSKU-1457-A"}],
            },
            "suggest_info": {
                "fba_available_sale_days": 100,
                "available_sale_days_fba": 20,
                "out_stock_date": "2026-05-12",
            },
            "data": {
                "amazon_quantity_info": {
                    "amazon_quantity_valid": 10,
                    "amazon_quantity_shipping": 5,
                    "afn_fulfillable_quantity": 8,
                    "reserved_fc_transfers": 1,
                    "reserved_fc_processing": 1,
                }
            },
            "ext_info": {"restock_status": 0},
        }

        record = classify_record(item, date(2026, 4, 7), {"1457": "Libraton JP-JP"}, {"1457"})

        self.assertIsNotNone(record)
        assert record is not None
        self.assertEqual(record.level, "A")
        self.assertEqual(record.reasons, ["断货时间(天数)=35天"])

    def test_classify_record_does_not_create_b_level_for_libraton_jp_jp(self) -> None:
        item = {
            "basic_info": {
                "asin": "B1457B",
                "hash_id": "hash-1457-b",
                "sid": "1457",
                "node_type": 1,
                "msku_fnsku_list": [{"msku": "MSKU-1457-B"}],
            },
            "suggest_info": {
                "fba_available_sale_days": 70,
                "available_sale_days_fba": 30,
                "out_stock_date": "2026-06-06",
            },
            "data": {
                "amazon_quantity_info": {
                    "amazon_quantity_valid": 10,
                    "amazon_quantity_shipping": 0,
                    "afn_fulfillable_quantity": 8,
                    "reserved_fc_transfers": 1,
                    "reserved_fc_processing": 1,
                }
            },
            "ext_info": {"restock_status": 0},
        }

        record = classify_record(item, date(2026, 4, 7), {"1457": "Libraton JP-JP"}, {"1457"})

        self.assertIsNone(record)

    def test_classify_record_uses_libraton_na_ca_a_level_thresholds(self) -> None:
        item = {
            "basic_info": {
                "asin": "B1444A",
                "hash_id": "hash-1444-a",
                "sid": "1444",
                "node_type": 1,
                "msku_fnsku_list": [{"msku": "MSKU-1444-A"}],
            },
            "suggest_info": {
                "fba_available_sale_days": 55,
                "available_sale_days_fba": 20,
                "out_stock_date": "2026-05-22",
            },
            "data": {
                "amazon_quantity_info": {
                    "amazon_quantity_valid": 10,
                    "amazon_quantity_shipping": 5,
                    "afn_fulfillable_quantity": 8,
                    "reserved_fc_transfers": 1,
                    "reserved_fc_processing": 1,
                }
            },
            "ext_info": {"restock_status": 0},
        }

        record = classify_record(item, date(2026, 4, 7), {"1444": "Libraton NA-CA"}, {"1444"})

        self.assertIsNotNone(record)
        assert record is not None
        self.assertEqual(record.level, "A")
        self.assertEqual(record.reasons, ["可售天数(FBA+在途)=55天", "断货时间(天数)=45天"])

    def test_classify_record_uses_libraton_na_ca_b_level_thresholds(self) -> None:
        item = {
            "basic_info": {
                "asin": "B1444B",
                "hash_id": "hash-1444-b",
                "sid": "1444",
                "node_type": 1,
                "msku_fnsku_list": [{"msku": "MSKU-1444-B"}],
            },
            "suggest_info": {
                "fba_available_sale_days": 70,
                "available_sale_days_fba": 60,
                "out_stock_date": "2026-06-06",
            },
            "data": {
                "amazon_quantity_info": {
                    "amazon_quantity_valid": 10,
                    "amazon_quantity_shipping": 5,
                    "afn_fulfillable_quantity": 8,
                    "reserved_fc_transfers": 1,
                    "reserved_fc_processing": 1,
                }
            },
            "ext_info": {"restock_status": 0},
        }

        record = classify_record(item, date(2026, 4, 7), {"1444": "Libraton NA-CA"}, {"1444"})

        self.assertIsNotNone(record)
        assert record is not None
        self.assertEqual(record.level, "B")
        self.assertEqual(record.reasons, ["可售天数(FBA)=断货时间(天数)=60天", "可售天数(FBA+在途)=70天"])

    def test_classify_record_does_not_alert_libraton_na_us_when_fba_plus_days_is_66(self) -> None:
        item = {
            "basic_info": {
                "asin": "B1443B",
                "hash_id": "hash-1443-b",
                "sid": "1443",
                "node_type": 1,
                "msku_fnsku_list": [{"msku": "MSKU-1443-B"}],
            },
            "suggest_info": {
                "fba_available_sale_days": 66,
                "available_sale_days_fba": 44,
                "out_stock_date": "2026-06-26",
            },
            "data": {
                "amazon_quantity_info": {
                    "amazon_quantity_valid": 10,
                    "amazon_quantity_shipping": 5,
                    "afn_fulfillable_quantity": 8,
                    "reserved_fc_transfers": 1,
                    "reserved_fc_processing": 1,
                }
            },
            "ext_info": {"restock_status": 0},
        }

        record = classify_record(item, date(2026, 4, 21), {"1443": "Libraton NA-US"}, {"1443"})

        self.assertIsNone(record)

    def test_classify_record_marks_a_level_when_fba_days_is_14(self) -> None:
        item = {
            "basic_info": {
                "asin": "B014",
                "hash_id": "hash-a-14",
                "sid": "1448",
                "node_type": 1,
                "msku_fnsku_list": [{"msku": "MSKU-A-14"}],
            },
            "suggest_info": {
                "fba_available_sale_days": 120,
                "available_sale_days_fba": 14,
                "out_stock_date": "2026-07-20",
            },
            "sales_info": {
                "sales_avg_7": 10,
                "sales_avg_14": 20,
                "sales_avg_30": 30,
            },
            "data": {
                "amazon_quantity_info": {
                    "amazon_quantity_valid": 10,
                    "amazon_quantity_shipping": 0,
                    "afn_fulfillable_quantity": 8,
                    "reserved_fc_transfers": 1,
                    "reserved_fc_processing": 1,
                }
            },
            "ext_info": {"restock_status": 0},
        }

        record = classify_record(item, date(2026, 4, 7), {"1448": "店铺A"}, {"1448"})

        self.assertIsNotNone(record)
        assert record is not None
        self.assertEqual(record.level, "A")
        self.assertIn("可售天数(FBA)=14天", record.reasons)

    def test_classify_record_marks_b_level_when_fba_days_is_30(self) -> None:
        item = {
            "basic_info": {
                "asin": "B030",
                "hash_id": "hash-b-30",
                "sid": "1448",
                "node_type": 1,
                "msku_fnsku_list": [{"msku": "MSKU-B-30"}],
            },
            "suggest_info": {
                "fba_available_sale_days": 120,
                "available_sale_days_fba": 30,
                "out_stock_date": "2026-07-20",
            },
            "sales_info": {
                "sales_avg_7": 10,
                "sales_avg_14": 20,
                "sales_avg_30": 30,
            },
            "data": {
                "amazon_quantity_info": {
                    "amazon_quantity_valid": 10,
                    "amazon_quantity_shipping": 0,
                    "afn_fulfillable_quantity": 8,
                    "reserved_fc_transfers": 1,
                    "reserved_fc_processing": 1,
                }
            },
            "ext_info": {"restock_status": 0},
        }

        record = classify_record(item, date(2026, 4, 7), {"1448": "店铺A"}, {"1448"})

        self.assertIsNotNone(record)
        assert record is not None
        self.assertEqual(record.level, "B")
        self.assertIn("可售天数(FBA)=30天", record.reasons)

    def test_classify_record_marks_c_level_when_fba_valid_is_zero_and_shipping_positive(self) -> None:
        item = {
            "basic_info": {
                "asin": "B001",
                "hash_id": "hash-c-1",
                "sid": "1448",
                "node_type": 1,
                "msku_fnsku_list": [{"msku": "MSKU-1"}],
            },
            "suggest_info": {
                "fba_available_sale_days": 120,
                "available_sale_days_fba": 0,
                "out_stock_date": "2026-08-20",
            },
            "sales_info": {
                "sales_avg_7": 10,
                "sales_avg_14": 20,
                "sales_avg_30": 30,
            },
            "data": {
                "amazon_quantity_info": {
                    "amazon_quantity_valid": 0,
                    "amazon_quantity_shipping": 18,
                    "afn_fulfillable_quantity": 11,
                    "reserved_fc_transfers": 4,
                    "reserved_fc_processing": 3,
                }
            },
            "ext_info": {"restock_status": 0},
        }

        record = classify_record(item, date(2026, 4, 7), {"1448": "店铺A"}, {"1448"})

        self.assertIsNotNone(record)
        assert record is not None
        self.assertEqual(record.level, "C")
        self.assertEqual(record.reasons, ["FBA库存=0 且 FBA在途=18"])
        self.assertEqual(record.fba_inventory, 0)
        self.assertEqual(record.fba_inbound_inventory, 18)
        self.assertEqual(record.fba_sellable_inventory, 11)
        self.assertEqual(record.fba_transfer_reserved_inventory, 4)
        self.assertEqual(record.fba_processing_inventory, 3)

    def test_aggregate_inventory_snapshot_sums_type_1_and_type_2_values(self) -> None:
        type_1_rows = [
            {
                "remark": {
                    "afn_fulfillable_quantity": "3",
                    "reserved_fc_transfers": "2",
                    "reserved_fc_processing": "1",
                }
            },
            {
                "remark": {
                    "afn_fulfillable_quantity": "4",
                    "reserved_fc_transfers": "5",
                    "reserved_fc_processing": "6",
                }
            },
        ]
        type_2_rows = [
            {"quantity": 7},
            {"quantity": "8"},
        ]

        snapshot = aggregate_inventory_snapshot(type_1_rows, type_2_rows)

        self.assertEqual(
            snapshot,
            InventorySnapshot(
                fba_sellable_inventory=7,
                fba_transfer_reserved_inventory=7,
                fba_processing_inventory=7,
                fba_inventory=21,
                fba_inbound_inventory=15,
            ),
        )

    def test_is_source_list_rate_limited_response_matches_api_error(self) -> None:
        self.assertTrue(
            is_source_list_rate_limited_response(
                {
                    "code": "3001008",
                    "msg": "new requests too frequently. please request later.",
                }
            )
        )
        self.assertFalse(is_source_list_rate_limited_response({"code": 0, "msg": "success"}))

    def test_parse_summary_items_can_create_c_level_from_source_list_values(self) -> None:
        item = {
            "basic_info": {
                "asin": "B001",
                "hash_id": "hash-c-2",
                "sid": "1448",
                "node_type": 1,
                "msku_fnsku_list": [{"msku": "MSKU-1"}],
            },
            "suggest_info": {
                "fba_available_sale_days": 120,
                "available_sale_days_fba": 31,
                "out_stock_date": "2026-08-20",
            },
            "sales_info": {
                "sales_avg_7": 10,
                "sales_avg_14": 20,
                "sales_avg_30": 30,
            },
            "data": {
                "amazon_quantity_info": {
                    "amazon_quantity_valid": 5,
                    "amazon_quantity_shipping": 0,
                    "afn_fulfillable_quantity": 5,
                    "reserved_fc_transfers": 0,
                    "reserved_fc_processing": 0,
                }
            },
            "ext_info": {"restock_status": 0},
        }

        records = parse_summary_items(
            [item],
            date(2026, 4, 7),
            {"1448": "店铺A"},
            ["1448"],
            {
                ("1448", "B001"): InventorySnapshot(
                    fba_sellable_inventory=0,
                    fba_transfer_reserved_inventory=0,
                    fba_processing_inventory=0,
                    fba_inventory=0,
                    fba_inbound_inventory=12,
                ),
            },
        )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0].level, "C")
        self.assertEqual(records[0].reasons, ["FBA库存=0 且 FBA在途=12"])
        self.assertEqual(records[0].fba_inventory, 0)
        self.assertEqual(records[0].fba_inbound_inventory, 12)

    def test_classify_record_computes_summary_daily_sales(self) -> None:
        item = make_summary_item(estimated_sale_avg_quantity=12.34)

        record = classify_record(item, date(2026, 4, 7), {"1448": "店铺A"}, {"1448"})

        self.assertIsNotNone(record)
        assert record is not None
        self.assertEqual(record.summary_daily_sales, 12.34)

    def test_build_report_rows_outputs_daily_sales_column(self) -> None:
        item = make_summary_item(estimated_sale_avg_quantity=12.34)

        record = classify_record(item, date(2026, 4, 7), {"1448": "店铺A"}, {"1448"})
        assert record is not None

        rows = build_report_rows([record])

        self.assertEqual(rows[0]["日均销量"], 12.34)
        self.assertEqual(rows[0]["FBA库存"], 5)
        self.assertEqual(rows[0]["FBA在途"], 12)
        self.assertEqual(rows[0]["FBA可售-可售"], 9)
        self.assertEqual(rows[0]["FBA可售-待调仓"], 2)
        self.assertEqual(rows[0]["FBA可售-调仓中"], 1)

    def test_build_listing_contact_map_groups_contacts_by_sid_and_asin(self) -> None:
        listing_rows = [
            {
                "sid": 1448,
                "asin": "B001",
                "principal_info": [
                    {"principal_name": "张三"},
                    {"principal_name": "李四"},
                ],
            },
            {
                "sid": "1448",
                "asin": "B001",
                "principal_info": [
                    {"principal_name": "李四"},
                    {"principal_name": "王五"},
                ],
            },
            {
                "sid": 1448,
                "asin": "B002",
                "principal_info": [],
            },
        ]

        contact_map = build_listing_contact_map(listing_rows)

        self.assertEqual(contact_map[("1448", "B001")], "张三, 李四, 王五")
        self.assertEqual(contact_map[("1448", "B002")], "")

    def test_build_report_rows_outputs_listing_contacts_column(self) -> None:
        item = make_summary_item()

        record = classify_record(item, date(2026, 4, 7), {"1448": "店铺A"}, {"1448"})
        assert record is not None
        record.listing_contacts = "张三, 李四"

        rows = build_report_rows([record])

        self.assertEqual(rows[0]["Listing联系人"], "张三, 李四")

    def test_export_alert_report_outputs_requested_column_order(self) -> None:
        item = make_summary_item(hash_id="hash-header-order")

        record = classify_record(item, date(2026, 4, 7), {"1448": "店铺A"}, {"1448"})
        assert record is not None

        with TemporaryDirectory() as tmp_dir:
            report_path = export_alert_report([record], date(2026, 4, 7), tmp_dir)
            workbook = load_workbook(Path(report_path))
            worksheet = workbook.active
            headers = [cell.value for cell in worksheet[1]]

        self.assertEqual(headers, REPORT_HEADERS)

    def test_export_alert_report_writes_main_report_into_date_directory(self) -> None:
        item = make_summary_item(hash_id="hash-main-date-dir")

        record = classify_record(item, date(2026, 4, 7), {"1448": "店铺A"}, {"1448"})
        assert record is not None

        with TemporaryDirectory() as tmp_dir:
            report_path = export_alert_report([record], date(2026, 4, 7), tmp_dir)

            self.assertEqual(
                Path(report_path),
                Path(tmp_dir) / "2026-04-07" / "LIBRATON库存预警-20260407.xlsx",
            )
            self.assertTrue(Path(report_path).exists())

    def test_export_alert_report_writes_store_specific_reports(self) -> None:
        item_a = make_summary_item(hash_id="hash-store-a", msku="MSKU-A")
        item_b = make_summary_item(asin="B002", hash_id="hash-store-b", sid="1446", msku="MSKU-B")

        record_a = classify_record(item_a, date(2026, 4, 7), {"1448": "店铺A"}, {"1448"})
        record_b = classify_record(item_b, date(2026, 4, 7), {"1446": "店铺B"}, {"1446"})
        assert record_a is not None
        assert record_b is not None

        with TemporaryDirectory() as tmp_dir:
            export_alert_report([record_a, record_b], date(2026, 4, 7), tmp_dir)

            store_a_path = Path(tmp_dir) / "2026-04-07" / "店铺A" / "LIBRATON库存预警-店铺A-20260407.xlsx"
            store_b_path = Path(tmp_dir) / "2026-04-07" / "店铺B" / "LIBRATON库存预警-店铺B-20260407.xlsx"

            self.assertTrue(store_a_path.exists())
            self.assertTrue(store_b_path.exists())

    def test_export_alert_report_merges_eu_de_and_eu_uk_into_one_eu_store_report(self) -> None:
        item_de = make_summary_item(hash_id="hash-eu-de", sid="1448", msku="MSKU-DE")
        item_uk = make_summary_item(asin="B002", hash_id="hash-eu-uk", sid="1446", msku="MSKU-UK")

        record_de = classify_record(item_de, date(2026, 4, 7), {"1448": "Libraton EU-DE"}, {"1448"})
        record_uk = classify_record(item_uk, date(2026, 4, 7), {"1446": "Libraton EU-UK"}, {"1446"})
        assert record_de is not None
        assert record_uk is not None

        with TemporaryDirectory() as tmp_dir:
            export_alert_report([record_de, record_uk], date(2026, 4, 7), tmp_dir)

            eu_path = Path(tmp_dir) / "2026-04-07" / "Libraton EU" / "LIBRATON库存预警-EU-20260407.xlsx"
            de_path = Path(tmp_dir) / "2026-04-07" / "Libraton EU-DE" / "LIBRATON库存预警-Libraton EU-DE-20260407.xlsx"
            uk_path = Path(tmp_dir) / "2026-04-07" / "Libraton EU-UK" / "LIBRATON库存预警-Libraton EU-UK-20260407.xlsx"

            self.assertTrue(eu_path.exists())
            self.assertFalse(de_path.exists())
            self.assertFalse(uk_path.exists())

    def test_export_scoped_alert_report_writes_only_merged_eu_report(self) -> None:
        item_de = make_summary_item(hash_id="hash-scoped-eu-de", sid="1448", msku="MSKU-DE")
        item_uk = make_summary_item(asin="B002", hash_id="hash-scoped-eu-uk", sid="1446", msku="MSKU-UK")

        record_de = classify_record(item_de, date(2026, 4, 21), {"1448": "Libraton EU-DE"}, {"1448"})
        record_uk = classify_record(item_uk, date(2026, 4, 21), {"1446": "Libraton EU-UK"}, {"1446"})
        assert record_de is not None
        assert record_uk is not None

        with TemporaryDirectory() as tmp_dir:
            report_path = export_scoped_alert_report([record_de, record_uk], date(2026, 4, 21), "Libraton EU", tmp_dir)

            self.assertEqual(
                Path(report_path),
                Path(tmp_dir) / "2026-04-21" / "Libraton EU" / "LIBRATON库存预警-EU-20260421.xlsx",
            )
            self.assertTrue(Path(report_path).exists())
            self.assertFalse((Path(tmp_dir) / "2026-04-21" / "LIBRATON库存预警-20260421.xlsx").exists())

    def test_export_alert_report_writes_one_sheet_per_store_in_main_report(self) -> None:
        item_a = make_summary_item(hash_id="hash-sheet-a", msku="MSKU-A")
        item_b = make_summary_item(asin="B002", hash_id="hash-sheet-b", sid="1446", msku="MSKU-B")

        record_a = classify_record(item_a, date(2026, 4, 7), {"1448": "店铺A"}, {"1448"})
        record_b = classify_record(item_b, date(2026, 4, 7), {"1446": "店铺B"}, {"1446"})
        assert record_a is not None
        assert record_b is not None

        with TemporaryDirectory() as tmp_dir:
            report_path = export_alert_report([record_a, record_b], date(2026, 4, 7), tmp_dir)
            workbook = load_workbook(Path(report_path))

        self.assertEqual(workbook.sheetnames, ["店铺A", "店铺B"])
        self.assertEqual(workbook["店铺A"]["A2"].value, "店铺A")
        self.assertEqual(workbook["店铺B"]["A2"].value, "店铺B")


if __name__ == "__main__":
    unittest.main()
