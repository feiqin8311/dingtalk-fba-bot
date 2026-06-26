#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .models import AlertRecord
from .store_policies import resolve_store_report_group_name

REPORT_HEADERS = [
    "店铺",
    "等级",
    "MSKU",
    "Listing联系人",
    "命中条数",
    "命中规则",
    "日均销量",
    "FBA库存",
    "可售天数(FBA)",
    "FBA在途",
    "可售天数(FBA+在途)",
    "断货时间",
    "断货天数",
    "FBA可售-可售",
    "FBA可售-待调仓",
    "FBA可售-调仓中",
]
REPORT_COLUMN_WIDTHS = {
    "A": 22,
    "B": 8,
    "C": 18,
    "D": 18,
    "E": 10,
    "F": 48,
    "G": 12,
    "H": 12,
    "I": 12,
    "J": 14,
    "K": 16,
    "L": 16,
    "M": 18,
    "N": 16,
    "O": 14,
    "P": 12,
}


def build_report_rows(alerts: list[AlertRecord]) -> list[dict]:
    rows: list[dict] = []
    seen: set[tuple[str, str, str]] = set()
    for item in alerts:
        reason_text = "；".join(item.reasons)
        for msku in item.mskus:
            key = (item.sid, item.level, msku)
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                {
                    "店铺": item.seller_name,
                    "等级": item.level,
                    "MSKU": msku,
                    "Listing联系人": item.listing_contacts,
                    "命中条数": len(item.reasons),
                    "命中规则": reason_text,
                    "日均销量": item.summary_daily_sales,
                    "FBA库存": item.fba_inventory,
                    "可售天数(FBA)": item.fba_days,
                    "FBA在途": item.fba_inbound_inventory,
                    "可售天数(FBA+在途)": item.fba_plus_days,
                    "断货时间": item.out_stock_date or "",
                    "断货天数": item.out_stock_days,
                    "FBA可售-可售": item.fba_sellable_inventory,
                    "FBA可售-待调仓": item.fba_transfer_reserved_inventory,
                    "FBA可售-调仓中": item.fba_processing_inventory,
                }
            )
    rows.sort(key=lambda row: (row["店铺"], row["等级"], row["断货天数"] if row["断货天数"] > 0 else 10**9, row["MSKU"]))
    return rows


def build_date_report_dir(today: date, output_dir: str) -> Path:
    return Path(output_dir) / today.isoformat()


def build_main_report_path(today: date, output_dir: str, report_name: str = "LIBRATON库存预警") -> Path:
    return build_date_report_dir(today, output_dir) / f"{report_name}-{today.strftime('%Y%m%d')}.xlsx"


def build_store_report_file_label(store_name: str) -> str:
    if store_name.startswith("Libraton "):
        return store_name[len("Libraton ") :]
    return store_name


def build_store_report_path(store_name: str, today: date, output_dir: str) -> Path:
    file_label = build_store_report_file_label(store_name)
    return build_date_report_dir(today, output_dir) / store_name / f"LIBRATON库存预警-{file_label}-{today.strftime('%Y%m%d')}.xlsx"


def group_rows_by_store(rows: list[dict]) -> dict[str, list[dict]]:
    rows_by_store: dict[str, list[dict]] = {}
    for row in rows:
        rows_by_store.setdefault(str(row["店铺"]), []).append(row)
    return rows_by_store


def group_rows_by_store_report(rows: list[dict]) -> dict[str, list[dict]]:
    rows_by_store_report: dict[str, list[dict]] = {}
    for row in rows:
        group_name = resolve_store_report_group_name(str(row["店铺"]))
        rows_by_store_report.setdefault(group_name, []).append(row)
    return rows_by_store_report


def write_report_workbook(rows: list[dict], file_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "库存预警"
    populate_report_worksheet(worksheet, rows)
    workbook.save(file_path)


def populate_report_worksheet(worksheet: object, rows: list[dict]) -> None:
    worksheet.append(REPORT_HEADERS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        worksheet.append([row[header] for header in REPORT_HEADERS])

    for column, width in REPORT_COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_multi_sheet_report_workbook(rows_by_store: dict[str, list[dict]], file_path: Path) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for store_name, store_rows in rows_by_store.items():
        worksheet = workbook.create_sheet(title=store_name[:31] or "库存预警")
        populate_report_worksheet(worksheet, store_rows)

    workbook.save(file_path)


def export_alert_report(
    alerts: list[AlertRecord],
    today: date,
    output_dir: str = "reports",
    *,
    include_store_reports: bool = True,
    main_report_name: str = "LIBRATON库存预警",
) -> str:
    rows = build_report_rows(alerts)
    rows_by_store = group_rows_by_store(rows)
    date_dir = build_date_report_dir(today, output_dir)
    date_dir.mkdir(parents=True, exist_ok=True)
    file_path = build_main_report_path(today, output_dir, main_report_name)
    write_multi_sheet_report_workbook(rows_by_store, file_path)

    if include_store_reports:
        rows_by_store_report = group_rows_by_store_report(rows)
        for store_name, store_rows in rows_by_store_report.items():
            store_file_path = build_store_report_path(store_name, today, output_dir)
            store_file_path.parent.mkdir(parents=True, exist_ok=True)
            write_report_workbook(store_rows, store_file_path)

    return str(file_path)


def export_scoped_alert_report(
    alerts: list[AlertRecord],
    today: date,
    store_name: str,
    output_dir: str = "reports",
) -> str:
    rows = build_report_rows(alerts)
    rows_by_store_report = group_rows_by_store_report(rows)
    store_rows = rows_by_store_report.get(store_name, [])
    if not store_rows:
        raise RuntimeError(f"未找到范围分表数据: {store_name}")

    store_file_path = build_store_report_path(store_name, today, output_dir)
    store_file_path.parent.mkdir(parents=True, exist_ok=True)
    write_report_workbook(store_rows, store_file_path)
    return str(store_file_path)
